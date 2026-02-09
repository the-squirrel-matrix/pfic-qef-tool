"""
Excel workbook I/O for PFIC QEF Tool.

Provides a single Excel file as an alternative to multiple JSON/CSV files.
This is more user-friendly for non-technical users.

Requires: openpyxl (pip install openpyxl)
"""

from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Optional, Union

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .models import (
    Config, Lot, Transaction, TransactionType, AISData, UnderlyingPFIC,
    round_shares, round_money
)


def check_openpyxl():
    """Raise ImportError if openpyxl is not available."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel support. "
            "Install it with: pip install openpyxl"
        )


# Style constants
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
INSTRUCTION_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_template_workbook(output_path: Union[str, Path], tax_year: int = 2024):
    """
    Create a template Excel workbook for user input.
    
    The workbook contains sheets for:
    - Instructions
    - Config
    - Beginning_Lots
    - Transactions
    - AIS_Data
    """
    check_openpyxl()
    
    wb = Workbook()
    
    # =========================================================================
    # Instructions Sheet
    # =========================================================================
    ws = wb.active
    ws.title = "Instructions"
    
    instructions = [
        ("PFIC QEF Tax Tool - Input Workbook", None),
        ("", None),
        ("This workbook collects all the information needed to calculate your", None),
        ("QEF income and basis adjustments for PFIC holdings.", None),
        ("", None),
        ("STEPS:", None),
        ("1. Fill in the Config sheet with your tax year and fund information", None),
        ("2. Fill in Beginning_Lots if you owned shares before this tax year", None),
        ("   (Leave empty if this is your first year owning this PFIC)", None),
        ("3. Fill in Transactions with any buys/sells during the year", None),
        ("   (Leave empty if no activity during the year)", None),
        ("4. Fill in AIS_Data from your fund's Annual Information Statement", None),
        ("5. Save this file and run the tool with --excel flag", None),
        ("", None),
        ("NOTES:", None),
        ("- All monetary amounts should be in USD unless noted otherwise", None),
        ("- Dates should be in YYYY-MM-DD format (e.g., 2024-03-15)", None),
        ("- Shares can have up to 4 decimal places", None),
        ("- Yellow cells contain instructions - don't modify them", None),
        ("", None),
        ("For help, see the README.md file included with the tool.", None),
    ]
    
    for row_num, (text, _) in enumerate(instructions, 1):
        cell = ws.cell(row=row_num, column=1, value=text)
        if row_num == 1:
            cell.font = Font(bold=True, size=14)
        elif text.startswith("STEPS:") or text.startswith("NOTES:"):
            cell.font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 70
    
    # =========================================================================
    # Config Sheet
    # =========================================================================
    ws = wb.create_sheet("Config")
    
    config_data = [
        ("Setting", "Value", "Description"),
        ("tax_year", tax_year, "The calendar year being processed"),
        ("pfic_ticker", "XEQT", "Ticker symbol of your PFIC (e.g., XEQT, VEQT)"),
        ("pfic_name", "iShares Core Equity ETF Portfolio", "Full name of the fund"),
        ("default_currency", "CAD", "Currency of your transactions (CAD or USD)"),
    ]
    
    for row_num, (setting, value, desc) in enumerate(config_data, 1):
        ws.cell(row=row_num, column=1, value=setting)
        ws.cell(row=row_num, column=2, value=value)
        ws.cell(row=row_num, column=3, value=desc)
        
        if row_num == 1:
            for col in range(1, 4):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
        else:
            ws.cell(row=row_num, column=3).fill = INSTRUCTION_FILL
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 50
    
    # =========================================================================
    # Beginning_Lots Sheet
    # =========================================================================
    ws = wb.create_sheet("Beginning_Lots")
    
    headers = ["lot_id", "ticker", "purchase_date", "quantity", "cost_basis_usd", "original_lot_id"]
    descriptions = [
        "Unique lot identifier (e.g., LOT-001)",
        "Fund ticker (e.g., XEQT) - must match Config",
        "Date purchased (YYYY-MM-DD)",
        "Number of shares/units (up to 4 decimals)",
        "Cost basis in USD (adjusted from prior years)",
        "Leave blank unless this lot was split"
    ]
    
    # Header row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    
    # Description row
    for col, desc in enumerate(descriptions, 1):
        cell = ws.cell(row=2, column=col, value=desc)
        cell.fill = INSTRUCTION_FILL
        cell.font = Font(italic=True, size=9)
    
    # Example row
    example = ["LOT-001", "XEQT", "2023-03-15", 100.0, 2500.00, ""]
    for col, val in enumerate(example, 1):
        ws.cell(row=3, column=col, value=val)
    
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # =========================================================================
    # Transactions Sheet
    # =========================================================================
    ws = wb.create_sheet("Transactions")
    
    headers = ["date", "type", "ticker", "quantity", "amount", "fees", "currency", "exchange_rate"]
    descriptions = [
        "Transaction date (YYYY-MM-DD preferred)",
        "Buy or Sell (case-insensitive)",
        "Fund ticker (e.g., XEQT)",
        "Number of shares/units",
        "Amount (excl. fees)",
        "Fees paid",
        "CAD or USD",
        "CAD→USD rate (optional if using BoC)"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    
    for col, desc in enumerate(descriptions, 1):
        cell = ws.cell(row=2, column=col, value=desc)
        cell.fill = INSTRUCTION_FILL
        cell.font = Font(italic=True, size=9)
    
    # Example rows
    examples = [
        ["2024-02-15", "BUY", "XEQT", 25.0, 650.00, 9.99, "CAD", 0.7450],
        ["2024-08-20", "SELL", "XEQT", 40.0, 1200.00, 9.99, "CAD", 0.7380],
    ]
    for row_num, example in enumerate(examples, 3):
        for col, val in enumerate(example, 1):
            ws.cell(row=row_num, column=col, value=val)
    
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # =========================================================================
    # AIS_Data Sheet
    # =========================================================================
    ws = wb.create_sheet("AIS_Data")
    
    # Main fund section
    ws.cell(row=1, column=1, value="TOP-LEVEL FUND (the fund you directly own)")
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.merge_cells('A1:C1')
    
    ais_fields = [
        ("Field", "Value", "Description"),
        ("fund_ticker", "XEQT", "Ticker symbol"),
        ("fund_name", "iShares Core Equity ETF Portfolio", "Full fund name"),
        ("ordinary_earnings_per_day_per_share_usd", "0.0003080775", "From AIS - per share per day"),
        ("net_capital_gains_per_day_per_share_usd", "0.0004661617", "From AIS - per share per day"),
        ("total_distributions_per_share_usd", "0.4498954722", "Total distributions for ENTIRE year"),
    ]
    
    for row_num, (field, value, desc) in enumerate(ais_fields, 2):
        ws.cell(row=row_num, column=1, value=field)
        ws.cell(row=row_num, column=2, value=value)
        ws.cell(row=row_num, column=3, value=desc)
        
        if row_num == 2:
            for col in range(1, 4):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
        else:
            ws.cell(row=row_num, column=3).fill = INSTRUCTION_FILL
    
    # Underlying PFICs section
    start_row = 10
    ws.cell(row=start_row, column=1, value="UNDERLYING PFICs (funds held by your fund)")
    ws.cell(row=start_row, column=1).font = Font(bold=True)
    ws.merge_cells(f'A{start_row}:D{start_row}')
    
    ws.cell(row=start_row + 1, column=1, value="(Add one row per underlying PFIC. Delete example rows if not applicable.)")
    ws.cell(row=start_row + 1, column=1).fill = INSTRUCTION_FILL
    ws.cell(row=start_row + 1, column=1).font = Font(italic=True)
    ws.merge_cells(f'A{start_row + 1}:D{start_row + 1}')
    
    underlying_headers = ["fund_ticker", "fund_name", "ordinary_earnings_per_day_per_share_usd", "net_capital_gains_per_day_per_share_usd"]
    for col, header in enumerate(underlying_headers, 1):
        cell = ws.cell(row=start_row + 2, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    
    # Example underlying PFICs
    underlying_examples = [
        ["XIC", "iShares Core S&P/TSX Capped Composite Index ETF", "0.0004731653", "0.0008148535"],
        ["XEF", "iShares Core MSCI EAFE IMI Index ETF", "0.0004135184", "0.0001008374"],
        ["XEC", "iShares Core MSCI Emerging Markets IMI Index ETF", "0.0000766569", "0.0000164945"],
    ]
    for row_num, example in enumerate(underlying_examples, start_row + 3):
        for col, val in enumerate(example, 1):
            ws.cell(row=row_num, column=col, value=val)
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 40
    
    # Save
    wb.save(output_path)
    return output_path


def load_from_excel(excel_path: Union[str, Path]) -> tuple[Config, list[Lot], list[Transaction], AISData]:
    """
    Load all input data from a single Excel workbook.
    
    Returns: (config, beginning_lots, transactions, ais_data)
    """
    check_openpyxl()
    
    wb = load_workbook(excel_path, data_only=True)
    
    # =========================================================================
    # Load Config
    # =========================================================================
    ws = wb["Config"]
    config_dict = {}
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] and row[1]:
            config_dict[row[0]] = row[1]
    
    config = Config(
        tax_year=int(config_dict.get("tax_year", 2024)),
        pfic_ticker=str(config_dict.get("pfic_ticker", "")),
        pfic_name=str(config_dict.get("pfic_name", "")),
        default_currency=str(config_dict.get("default_currency", "CAD")).upper(),
    )
    
    # =========================================================================
    # Load Beginning Lots
    # =========================================================================
    ws = wb["Beginning_Lots"]
    lots = []
    
    # Columns: lot_id, ticker, purchase_date, shares, cost_basis_usd, original_lot_id
    for row in ws.iter_rows(min_row=3, values_only=True):  # Skip header and description
        lot_id = str(row[0]) if row[0] else None
        if not lot_id:
            continue
        
        # Get ticker (column 1) - filter to match config
        ticker = str(row[1]).upper() if row[1] else ""
        if ticker and config.pfic_ticker and ticker != config.pfic_ticker.upper():
            continue  # Skip lots for different PFICs
            
        purchase_date = row[2]
        if isinstance(purchase_date, str):
            purchase_date = date.fromisoformat(purchase_date)
        elif hasattr(purchase_date, 'date'):
            purchase_date = purchase_date.date()
        
        lots.append(Lot(
            lot_id=lot_id,
            ticker=ticker or config.pfic_ticker,
            purchase_date=purchase_date,
            shares=Decimal(str(row[3])) if row[3] else Decimal("0"),
            cost_basis_usd=Decimal(str(row[4])) if row[4] else Decimal("0"),
            original_lot_id=str(row[5]) if row[5] else None,
        ))
    
    # =========================================================================
    # Load Transactions
    # =========================================================================
    ws = wb["Transactions"]
    transactions = []
    
    # Columns: date, type, ticker, shares, amount, commission, currency, exchange_rate
    for row in ws.iter_rows(min_row=3, values_only=True):  # Skip header and description
        if not row[0] or not row[1]:
            continue
        
        txn_date = row[0]
        if isinstance(txn_date, str):
            txn_date = date.fromisoformat(txn_date)
        elif hasattr(txn_date, 'date'):
            txn_date = txn_date.date()
        
        txn_type_str = str(row[1]).upper()
        if txn_type_str == "BUY":
            txn_type = TransactionType.BUY
        elif txn_type_str == "SELL":
            txn_type = TransactionType.SELL
        else:
            continue
        
        # Get ticker (column 2) - filter to match config
        ticker = str(row[2]).upper() if row[2] else ""
        if ticker and config.pfic_ticker and ticker != config.pfic_ticker.upper():
            continue  # Skip transactions for different PFICs
        
        currency = str(row[6]).upper() if row[6] else config.default_currency
        
        # Parse optional exchange rate (column 7)
        exchange_rate = None
        if len(row) > 7 and row[7]:
            exchange_rate = Decimal(str(row[7]))
        
        transactions.append(Transaction(
            date=txn_date,
            transaction_type=txn_type,
            ticker=ticker or config.pfic_ticker,
            shares=Decimal(str(row[3])) if row[3] else Decimal("0"),
            amount=Decimal(str(row[4])) if row[4] else Decimal("0"),
            commission=Decimal(str(row[5])) if row[5] else Decimal("0"),
            currency=currency,
            exchange_rate=exchange_rate,
        ))
    
    # =========================================================================
    # Load AIS Data
    # =========================================================================
    ws = wb["AIS_Data"]
    
    ais_dict = {}
    for row in ws.iter_rows(min_row=3, max_row=8, max_col=2, values_only=True):
        if row[0] and row[1]:
            ais_dict[row[0]] = row[1]
    
    # Load underlying PFICs
    underlying_pfics = []
    for row in ws.iter_rows(min_row=13, values_only=True):  # Start after underlying header
        if not row[0] or not row[1]:
            continue
        underlying_pfics.append(UnderlyingPFIC(
            fund_ticker=str(row[0]),
            fund_name=str(row[1]),
            ordinary_earnings_per_day_per_share_usd=Decimal(str(row[2])) if row[2] else Decimal("0"),
            net_capital_gains_per_day_per_share_usd=Decimal(str(row[3])) if row[3] else Decimal("0"),
        ))
    
    ais_data = AISData(
        tax_year=int(ais_dict.get("tax_year", config.tax_year)),
        fund_ticker=str(ais_dict.get("fund_ticker", config.pfic_ticker)),
        fund_name=str(ais_dict.get("fund_name", config.pfic_name)),
        ordinary_earnings_per_day_per_share_usd=Decimal(str(ais_dict.get("ordinary_earnings_per_day_per_share_usd", "0"))),
        net_capital_gains_per_day_per_share_usd=Decimal(str(ais_dict.get("net_capital_gains_per_day_per_share_usd", "0"))),
        total_distributions_per_share_usd=Decimal(str(ais_dict.get("total_distributions_per_share_usd", "0"))),
        underlying_pfics=underlying_pfics,
    )
    
    return config, lots, transactions, ais_data


def save_results_to_excel(
    output_path: Union[str, Path],
    form_8621_data: list,
    sales: list,
    adjustments: list,
    ending_lots: list,
    report,
):
    """
    Save all results to a single Excel workbook.
    
    Creates sheets for:
    - Summary
    - Form_8621_Data
    - Sales_Report
    - Basis_Adjustments
    - Year_End_Lots
    """
    check_openpyxl()
    
    wb = Workbook()
    
    # =========================================================================
    # Summary Sheet
    # =========================================================================
    ws = wb.active
    ws.title = "Summary"
    
    ws.cell(row=1, column=1, value=f"PFIC QEF Tax Report - {report.tax_year}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    ws.cell(row=2, column=1, value=f"{report.pfic_name} ({report.pfic_ticker})")
    
    summary_data = [
        ("", ""),
        ("Beginning Lots", len(report.beginning_lots)),
        ("Transactions Processed", len(report.transactions_processed)),
        ("Lots Sold", len(report.lots_sold)),
        ("Ending Lots", len(report.ending_lots)),
        ("", ""),
        ("Total QEF Ordinary Earnings", sum(f.line_6a_ordinary_earnings_usd for f in form_8621_data)),
        ("Total QEF Capital Gains", sum(f.line_7a_net_capital_gains_usd for f in form_8621_data)),
        ("Form 8621 Required", len(form_8621_data)),
    ]
    
    for row_num, (label, value) in enumerate(summary_data, 4):
        ws.cell(row=row_num, column=1, value=label)
        if isinstance(value, Decimal):
            ws.cell(row=row_num, column=2, value=float(value))
            ws.cell(row=row_num, column=2).number_format = '$#,##0.00'
        else:
            ws.cell(row=row_num, column=2, value=value)
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    # =========================================================================
    # Form_8621_Data Sheet
    # =========================================================================
    ws = wb.create_sheet("Form_8621_Data")
    
    headers = ["Fund Ticker", "Fund Name", "Direct Holding", "Line 6a Ordinary", "Line 7a Cap Gains", "Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    
    for row_num, f in enumerate(form_8621_data, 2):
        ws.cell(row=row_num, column=1, value=f.fund_ticker)
        ws.cell(row=row_num, column=2, value=f.fund_name)
        ws.cell(row=row_num, column=3, value="Yes" if f.is_direct_holding else "No")
        ws.cell(row=row_num, column=4, value=float(f.line_6a_ordinary_earnings_usd))
        ws.cell(row=row_num, column=5, value=float(f.line_7a_net_capital_gains_usd))
        ws.cell(row=row_num, column=6, value=float(f.line_6a_ordinary_earnings_usd + f.line_7a_net_capital_gains_usd))
        
        for col in range(4, 7):
            ws.cell(row=row_num, column=col).number_format = '$#,##0.00'
    
    for col, width in enumerate([15, 45, 15, 18, 18, 15], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # =========================================================================
    # Sales_Report Sheet
    # =========================================================================
    if sales:
        ws = wb.create_sheet("Sales_Report")
        
        headers = ["Lot ID", "Purchase Date", "Sale Date", "Shares", "Adj. Basis", "Proceeds", "Gain/Loss", "Type"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        
        for row_num, s in enumerate(sales, 2):
            ws.cell(row=row_num, column=1, value=s.lot_id)
            ws.cell(row=row_num, column=2, value=s.purchase_date.isoformat())
            ws.cell(row=row_num, column=3, value=s.sale_date.isoformat())
            ws.cell(row=row_num, column=4, value=float(s.shares_sold))
            ws.cell(row=row_num, column=5, value=float(s.cost_basis_adjusted_usd))
            ws.cell(row=row_num, column=6, value=float(s.proceeds_usd))
            ws.cell(row=row_num, column=7, value=float(s.gain_loss_usd))
            ws.cell(row=row_num, column=8, value=s.gain_type.value)
            
            for col in range(5, 8):
                ws.cell(row=row_num, column=col).number_format = '$#,##0.00'
    
    # =========================================================================
    # Basis_Adjustments Sheet
    # =========================================================================
    ws = wb.create_sheet("Basis_Adjustments")
    
    headers = ["Lot ID", "Shares", "Days Held", "Ord. Earnings", "Cap. Gains", "Distributions", "Net Adj.", "New Basis"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    
    for row_num, a in enumerate(adjustments, 2):
        ws.cell(row=row_num, column=1, value=a.lot_id)
        ws.cell(row=row_num, column=2, value=float(a.shares))
        ws.cell(row=row_num, column=3, value=a.days_held_in_year)
        ws.cell(row=row_num, column=4, value=float(a.ordinary_earnings_usd))
        ws.cell(row=row_num, column=5, value=float(a.capital_gains_usd))
        ws.cell(row=row_num, column=6, value=float(a.distributions_usd))
        ws.cell(row=row_num, column=7, value=float(a.net_adjustment_usd))
        ws.cell(row=row_num, column=8, value=float(a.basis_after_usd))
        
        for col in range(4, 9):
            ws.cell(row=row_num, column=col).number_format = '$#,##0.00'
    
    # =========================================================================
    # Year_End_Lots Sheet
    # =========================================================================
    ws = wb.create_sheet("Year_End_Lots")
    
    ws.cell(row=1, column=1, value="Use this data as Beginning_Lots for next year")
    ws.cell(row=1, column=1).font = Font(bold=True, italic=True)
    ws.merge_cells('A1:E1')
    
    headers = ["lot_id", "purchase_date", "quantity", "cost_basis_usd", "original_lot_id"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    
    for row_num, lot in enumerate(ending_lots, 3):
        ws.cell(row=row_num, column=1, value=lot.lot_id)
        ws.cell(row=row_num, column=2, value=lot.ticker or "")
        ws.cell(row=row_num, column=3, value=lot.purchase_date.isoformat())
        ws.cell(row=row_num, column=4, value=float(lot.shares))
        ws.cell(row=row_num, column=5, value=float(lot.cost_basis_usd))
        ws.cell(row=row_num, column=6, value=lot.original_lot_id or "")
        
        ws.cell(row=row_num, column=4).number_format = '0.0000'
        ws.cell(row=row_num, column=5).number_format = '$#,##0.00'
    
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    wb.save(output_path)
    return output_path
